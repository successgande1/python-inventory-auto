#import the openpyxl module
import openpyxl

#Create an Inventory object File using the load_workbook method
inventory_file = openpyxl.load_workbook("inventory.xlsx")

#Load a worksheet in the Inventory file
products_list = inventory_file['Inventories']

#Create an empty dictionary of list of products per supplier
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}



#print(products_list.max_row)

#Loop through the list of products starting from the first to the last
for product_row in range(2, products_list.max_row + 1):
    #Create Variable for the supplier and locate the cell Number in the worksheet with suppliers' Names
    supplier_name = products_list.cell(product_row, 4).value
     #Create Variable for the Inventory and locate the cell Number in the worksheet with list of Inventories
    inventory = products_list.cell(product_row, 2).value
     #Create Variable for the Price and locate the cell Number in the worksheet with the List of Prices
    price = products_list.cell(product_row, 3).value
     #Create Variable for the Product Number and locate the cell Address in the worksheet with the List of Prices
    products_num = products_list.cell(product_row, 1).value
    #Create a variable for a New column using its cell address
    inventory_price = products_list.cell(product_row, 5)
   # inventory_price_heading = products_list[5]

    #Check if supplier Name in the Dictionary
    if supplier_name in products_per_supplier:
        #The First Supplier Found should get its Name from the Dictionary using the get function
        #using the supplier name
        current_num_supplier = products_per_supplier.get(supplier_name)
        #Get the Dictionary Key which is the Supplier Name and add its products
        products_per_supplier[supplier_name] = current_num_supplier + 1
    else:
        #print("Adding Supplier")
        products_per_supplier[supplier_name] = 1
    #Logic to calculate Total Value of products per Supplier
    if supplier_name in total_value_per_supplier:
        inventory_value_supplier = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = inventory_value_supplier + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    #Logic To Calculate Inventory Less Than 10
    if inventory < 10:
        
        products_under_10_inv[products_num] = inventory

    #Add New Column Heading
    products_list['E1'] = "Inventory Price"

    #Add new column for the Inventory total Price with corresponding values
    inventory_price.value = inventory * price

#Price products inventory under 10
print(products_under_10_inv)

#Price Total Value of Products Per Supplier
print(total_value_per_supplier)

#Price Number of Products Per Supplier
print(products_per_supplier)

#Save the Workbook with a new Name
inventory_file.save("inventory_with_total_price.xlsx")
